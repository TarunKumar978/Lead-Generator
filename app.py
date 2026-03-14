"""
Silasya & Shumitra — AI Lead Finder
Flask + MySQL Backend
"""

from flask import Flask, request, jsonify, render_template, session, Response
from flask_cors import CORS
import mysql.connector
from mysql.connector import pooling
import os
import json
import anthropic
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "silasya-secret-2025")
CORS(app)

# ─── DB Pool ─────────────────────────────────────────────────────────────────

db_config = {
    "host": os.getenv("DB_HOST", "localhost"),
    "port": int(os.getenv("DB_PORT", 3306)),
    "user": os.getenv("DB_USER", "root"),
    "password": os.getenv("DB_PASSWORD", ""),
    "database": os.getenv("DB_NAME", "silasya_leads"),
}

connection_pool = None


def get_pool():
    global connection_pool
    if connection_pool is None:
        connection_pool = pooling.MySQLConnectionPool(
            pool_name="silasya_pool",
            pool_size=5,
            **db_config
        )
    return connection_pool


def get_db():
    return get_pool().get_connection()


# ─── DB Init ─────────────────────────────────────────────────────────────────

def init_db():
    conn = mysql.connector.connect(
        host=db_config["host"],
        port=db_config["port"],
        user=db_config["user"],
        password=db_config["password"],
    )
    cursor = conn.cursor()
    cursor.execute(f"CREATE DATABASE IF NOT EXISTS `{db_config['database']}` CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci")
    cursor.execute(f"USE `{db_config['database']}`")
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS leads (
            id INT AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(255) NOT NULL,
            type VARCHAR(10) DEFAULT 'b2c',
            category VARCHAR(255),
            country VARCHAR(100),
            city VARCHAR(100),
            email VARCHAR(255),
            phone VARCHAR(100),
            website VARCHAR(500),
            instagram VARCHAR(255),
            linkedin VARCHAR(500),
            facebook VARCHAR(500),
            whatsapp VARCHAR(100),
            description TEXT,
            why_good TEXT,
            potential_value VARCHAR(100),
            tags JSON,
            score INT DEFAULT 50,
            status VARCHAR(50) DEFAULT 'new',
            source VARCHAR(255),
            notes TEXT,
            saved_by VARCHAR(100),
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS search_history (
            id INT AUTO_INCREMENT PRIMARY KEY,
            business VARCHAR(50),
            country VARCHAR(255),
            niche VARCHAR(255),
            lead_type VARCHAR(255),
            channels TEXT,
            keywords TEXT,
            leads_found INT DEFAULT 0,
            searched_by VARCHAR(100),
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS team_members (
            id INT AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(100) NOT NULL,
            email VARCHAR(255) UNIQUE NOT NULL,
            role VARCHAR(20) DEFAULT 'member',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS reminders (
            id INT AUTO_INCREMENT PRIMARY KEY,
            lead_id INT NOT NULL,
            note TEXT,
            remind_at DATETIME NOT NULL,
            done TINYINT DEFAULT 0,
            created_by VARCHAR(100),
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)
    if cursor.fetchone()[0] == 0:
        cursor.execute("INSERT INTO team_members (name, email, role) VALUES ('Admin', 'admin@silasya.com', 'admin')")
    conn.commit()
    cursor.close()
    conn.close()
    print("✅ Database initialized successfully")


# ─── Auth ────────────────────────────────────────────────────────────────────

APP_USERNAME = os.getenv("APP_USERNAME", "silasya")
APP_PASSWORD = os.getenv("APP_PASSWORD", "silasya2025")

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            if request.path.startswith("/api/"):
                return jsonify({"error": "Unauthorized"}), 401
            return render_template("login.html")
        return f(*args, **kwargs)
    return decorated

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        data = request.get_json(silent=True) or request.form
        if data.get("username") == APP_USERNAME and data.get("password") == APP_PASSWORD:
            session["logged_in"] = True
            session["user"] = data.get("username")
            return jsonify({"success": True})
        return jsonify({"error": "Wrong username or password"}), 401
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return render_template("login.html")


# ─── Routes ──────────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def index():
    return render_template("index.html")


@app.route("/api/status")
@login_required
def status():
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM leads")
        total = cursor.fetchone()[0]
        cursor.close()
        conn.close()
        return jsonify({"status": "ok", "total_leads": total, "db": "connected"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


# ─── AI Search ───────────────────────────────────────────────────────────────

@app.route("/api/search", methods=["POST"])
@login_required
def ai_search():
    try:
        data = request.get_json()
        business = data.get("business", "both")
        country = data.get("country", "India and worldwide")
        niche = data.get("niche", "organic products")
        lead_type = data.get("lead_type", "all")
        channels = data.get("channels", "Google, LinkedIn, Instagram")
        keywords = data.get("keywords", "")
        saved_by = data.get("saved_by", "Team")

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO search_history (business, country, niche, lead_type, channels, keywords, searched_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (business, country, niche, lead_type, channels, keywords, saved_by))
        conn.commit()
        search_id = cursor.lastrowid

        cursor.execute("""
            DELETE FROM search_history WHERE id NOT IN (
                SELECT id FROM (SELECT id FROM search_history ORDER BY created_at DESC LIMIT 100) AS tmp
            )
        """)
        conn.commit()
        cursor.close()
        conn.close()

        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

        prompt = f"""You are a lead generation expert for two Indian organic businesses:
- SILASYA: B2C organic apparel, toys, home decor brand
- SHUMITRA: B2B export arm selling organic products worldwide

Use web search to find REAL businesses, stores, importers and buyers actively looking for organic products.
Search Google, LinkedIn, Instagram, IndiaMART, Alibaba, Etsy and trade directories.

Search parameters:
Business Focus: {business}
Target Country/Region: {country}
Product Niche: {niche}
Lead Type: {lead_type}
Channels: {channels}
Keywords: {keywords}

Return ONLY a valid JSON object with this exact structure:
{{
  "leads": [
    {{
      "name": "Real company or person name found online",
      "type": "b2c or b2b",
      "category": "e.g. Organic Retailer, Wholesale Importer",
      "country": "country",
      "city": "city",
      "email": "real email if found",
      "phone": "real phone if found",
      "website": "real website URL",
      "instagram": "@real handle if found",
      "linkedin": "real LinkedIn URL if found",
      "whatsapp": "phone number",
      "description": "What this business does based on what you found online",
      "why_good": "Why they are a good lead for Silasya/Shumitra",
      "potential_value": "Estimated order value e.g. $2,000-$5,000/month",
      "score": 85,
      "tags": ["tag1", "tag2"],
      "source": "Where you found them e.g. IndiaMART, Instagram, Google",
      "verified": true,
      "demand_signals": "What products they are actively buying or searching for"
    }}
  ],
  "demand_intelligence": {{
    "market_summary": "2-3 sentences on current market demand for {niche} in {country}",
    "trending_products": ["product1", "product2", "product3"],
    "avg_order_value": "Typical order size in this market",
    "best_channels": ["channel1", "channel2"],
    "peak_season": "When demand is highest e.g. Oct-Dec",
    "buyer_pain_points": ["pain1", "pain2", "pain3"],
    "price_range": "Typical price buyers pay",
    "competition_level": "Low / Medium / High",
    "opportunity_score": 85,
    "how_to_convert": ["tip1", "tip2", "tip3"]
  }}
}}

Find at least 12 real leads using web search. Return ONLY the JSON object, no other text."""

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=6000,
            tools=[{"type": "web_search_20250305", "name": "web_search"}],
            messages=[{"role": "user", "content": prompt}]
        )

        # Extract text from all response blocks
        raw = ""
        for block in message.content:
            if hasattr(block, "text"):
                raw += block.text

        raw = raw.strip().replace("```json", "").replace("```", "").strip()
        start = raw.find("{")
        end = raw.rfind("}")
        if start != -1 and end != -1:
            raw = raw[start:end+1]

        parsed = json.loads(raw)
        leads = parsed.get("leads", [])
        demand_intelligence = parsed.get("demand_intelligence", {})

        # Normalize fields
        for lead in leads:
            if not lead.get("name"):
                lead["name"] = lead.pop("company_name", None) or lead.pop("business_name", None) or "Unknown Lead"
            lead["type"] = "b2b" if "b2b" in str(lead.get("type","")).lower() else "b2c"
            try:
                lead["score"] = int(lead.get("score", 50))
            except:
                lead["score"] = 50
            if not isinstance(lead.get("tags"), list):
                lead["tags"] = []

        # Update search history
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("UPDATE search_history SET leads_found = %s WHERE id = %s", (len(leads), search_id))
        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({"success": True, "leads": leads, "count": len(leads), "demand_intelligence": demand_intelligence})


    except json.JSONDecodeError as e:
        return jsonify({"error": f"AI returned invalid data: {str(e)}"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Save Lead ───────────────────────────────────────────────────────────────

@app.route("/api/leads", methods=["GET"])
@login_required
def get_leads():
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        status_filter = request.args.get("status", "")
        type_filter = request.args.get("type", "")
        search = request.args.get("search", "")

        query = "SELECT * FROM leads WHERE 1=1"
        params = []

        if status_filter:
            query += " AND status = %s"
            params.append(status_filter)
        if type_filter:
            query += " AND type = %s"
            params.append(type_filter)
        if search:
            query += " AND (name LIKE %s OR email LIKE %s OR country LIKE %s OR category LIKE %s)"
            s = f"%{search}%"
            params.extend([s, s, s, s])

        query += " ORDER BY created_at DESC"
        cursor.execute(query, params)
        leads = cursor.fetchall()

        for lead in leads:
            if lead.get("tags") and isinstance(lead["tags"], str):
                try:
                    lead["tags"] = json.loads(lead["tags"])
                except:
                    lead["tags"] = []
            if lead.get("created_at"):
                lead["created_at"] = lead["created_at"].isoformat()
            if lead.get("updated_at"):
                lead["updated_at"] = lead["updated_at"].isoformat()

        cursor.close()
        conn.close()
        return jsonify({"success": True, "leads": leads, "count": len(leads)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/leads", methods=["POST"])
@login_required
def save_lead():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No data provided"}), 400

        name = data.get("name") or "Unknown Lead"
        tags = data.get("tags", [])
        if isinstance(tags, list):
            tags = json.dumps(tags)

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO leads (name, type, category, country, city, email, phone, website,
                instagram, linkedin, facebook, whatsapp, description, why_good,
                potential_value, tags, score, status, source, saved_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            name,
            data.get("type", "b2c"),
            data.get("category", ""),
            data.get("country", ""),
            data.get("city", ""),
            data.get("email", ""),
            data.get("phone", ""),
            data.get("website", ""),
            data.get("instagram", ""),
            data.get("linkedin", ""),
            data.get("facebook", ""),
            data.get("whatsapp", ""),
            data.get("description", ""),
            data.get("why_good", ""),
            data.get("potential_value", ""),
            tags,
            int(data.get("score", 50)),
            data.get("status", "new"),
            data.get("source", "AI Search"),
            data.get("saved_by", "Team"),
        ))
        conn.commit()
        lead_id = cursor.lastrowid
        cursor.close()
        conn.close()

        return jsonify({"success": True, "id": lead_id, "message": f"Lead '{name}' saved!"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/leads/<int:lead_id>", methods=["PUT"])
@login_required
def update_lead(lead_id):
    try:
        data = request.get_json()
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE leads SET status=%s, notes=%s WHERE id=%s
        """, (data.get("status", "new"), data.get("notes", ""), lead_id))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/leads/<int:lead_id>", methods=["DELETE"])
@login_required
def delete_lead(lead_id):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM leads WHERE id=%s", (lead_id,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Stats ───────────────────────────────────────────────────────────────────

@app.route("/api/stats")
@login_required
def get_stats():
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT COUNT(*) as total FROM leads")
        total = cursor.fetchone()["total"]
        cursor.execute("SELECT COUNT(*) as hot FROM leads WHERE score >= 80")
        hot = cursor.fetchone()["hot"]
        cursor.execute("SELECT COUNT(*) as converted FROM leads WHERE status = 'converted'")
        converted = cursor.fetchone()["converted"]
        cursor.execute("SELECT COUNT(DISTINCT country) as countries FROM leads")
        countries = cursor.fetchone()["countries"]
        cursor.close()
        conn.close()
        return jsonify({"total": total, "hot": hot, "converted": converted, "countries": countries})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Export CSV ──────────────────────────────────────────────────────────────

@app.route("/api/export/csv")
@login_required
def export_csv():
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM leads ORDER BY created_at DESC")
        leads = cursor.fetchall()
        cursor.close()
        conn.close()

        lines = ["Name,Type,Category,Country,City,Email,Phone,Website,Score,Status,Saved By,Created At"]
        for l in leads:
            lines.append(f'"{l.get("name","")}", "{l.get("type","")}", "{l.get("category","")}", "{l.get("country","")}", "{l.get("city","")}", "{l.get("email","")}", "{l.get("phone","")}", "{l.get("website","")}", "{l.get("score","")}", "{l.get("status","")}", "{l.get("saved_by","")}", "{l.get("created_at","")}"')

        csv = "\n".join(lines)
        return Response(csv, mimetype="text/csv", headers={"Content-Disposition": "attachment;filename=silasya_leads.csv"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Team ────────────────────────────────────────────────────────────────────

@app.route("/api/team", methods=["GET"])
@login_required
def get_team():
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM team_members ORDER BY created_at")
        members = cursor.fetchall()
        for m in members:
            if m.get("created_at"):
                m["created_at"] = m["created_at"].isoformat()
        cursor.close()
        conn.close()
        return jsonify({"success": True, "members": members})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/team", methods=["POST"])
@login_required
def add_team():
    try:
        data = request.get_json()
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO team_members (name, email, role) VALUES (%s,%s,%s)
        """, (data.get("name"), data.get("email"), data.get("role", "member")))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── AI Outreach ─────────────────────────────────────────────────────────────

@app.route("/api/outreach/<int:lead_id>", methods=["POST"])
@login_required
def generate_outreach(lead_id):
    try:
        data = request.get_json()
        channel = data.get("channel", "email")

        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM leads WHERE id=%s", (lead_id,))
        lead = cursor.fetchone()
        cursor.close()
        conn.close()

        if not lead:
            return jsonify({"error": "Lead not found"}), 404

        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

        prompt = f"""Write a {channel} outreach message for this lead on behalf of Silasya & Shumitra (Indian organic brand).

Lead: {lead.get('name')}
Type: {lead.get('type')}
Category: {lead.get('category')}
Country: {lead.get('country')}
Description: {lead.get('description')}

Write a short, friendly, professional {channel} message. Keep it under 150 words."""

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=500,
            messages=[{"role": "user", "content": prompt}]
        )

        return jsonify({"success": True, "message": message.content[0].text})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Buyer Requirements ──────────────────────────────────────────────────────

@app.route("/api/buyer-requirements", methods=["POST"])
@login_required
def buyer_requirements():
    try:
        data = request.get_json()
        niche = data.get("niche", "organic products")
        country = data.get("country", "worldwide")
        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
        prompt = f"""Use web search to find REAL active buyer requirements and RFQs posted on:
IndiaMART, Alibaba, TradeIndia, Global Sources, ExportHub, EC21, Made-in-China, DHgate, Thomasnet, Fibre2Fashion, Tradewheels, Go4WorldBusiness, and any other B2B trade portals.

Search for buyers looking for: {niche}
Target regions: {country}

Return ONLY a valid JSON array of at least 10 real buyer requirements:
[
  {{
    "buyer_name": "Company or buyer name",
    "platform": "IndiaMART / Alibaba / TradeIndia / etc",
    "requirement": "Exactly what they need e.g. 500 organic cotton t-shirts",
    "quantity": "e.g. 500 units / 1000 kg",
    "budget": "e.g. $2,000-$5,000 or ₹1.5L",
    "country": "Buyer country",
    "city": "Buyer city",
    "timeline": "e.g. Within 30 days / Urgent",
    "contact": "email or phone if available",
    "posted": "e.g. 2 days ago / This week",
    "verified": true,
    "match_score": 85,
    "url": "Direct link to the requirement if available",
    "notes": "Any extra details about what they want"
  }}
]
Return ONLY the JSON array, no other text."""

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=5000,
            tools=[{"type": "web_search_20250305", "name": "web_search"}],
            messages=[{"role": "user", "content": prompt}]
        )
        raw = ""
        for block in message.content:
            if hasattr(block, "text"):
                raw += block.text
        raw = raw.strip().replace("```json", "").replace("```", "").strip()
        start = raw.find("[")
        end = raw.rfind("]")
        if start != -1 and end != -1:
            raw = raw[start:end+1]
        requirements = json.loads(raw)
        return jsonify({"success": True, "requirements": requirements, "count": len(requirements)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── AI Email Writer ──────────────────────────────────────────────────────────

@app.route("/api/email/<int:lead_id>", methods=["POST"])
@login_required
def write_email(lead_id):
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM leads WHERE id=%s", (lead_id,))
        lead = cursor.fetchone()
        cursor.close()
        conn.close()
        if not lead:
            return jsonify({"error": "Lead not found"}), 404

        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
        prompt = f"""Write a professional sales email for this lead on behalf of Silasya & Shumitra (Indian organic brand exporting worldwide).

Lead Details:
Name: {lead.get('name')}
Type: {lead.get('type','b2b')}
Category: {lead.get('category','')}
Country: {lead.get('country','')}
Description: {lead.get('description','')}
Why Good Lead: {lead.get('why_good','')}
Demand Signals: {lead.get('demand_signals','')}

Write 3 versions:
1. SHORT (50 words) — quick intro for cold email
2. MEDIUM (100 words) — warm pitch with value proposition  
3. DETAILED (200 words) — full pitch with product details, certifications, MOQ

Return ONLY a valid JSON object:
{{
  "subject": "Best email subject line",
  "short": "50 word version",
  "medium": "100 word version", 
  "detailed": "200 word version",
  "whatsapp": "Short 30 word WhatsApp message version",
  "follow_up": "Follow-up message to send after 3 days of no reply"
}}"""

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=2000,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = message.content[0].text.strip().replace("```json","").replace("```","").strip()
        start = raw.find("{")
        end = raw.rfind("}")
        if start != -1 and end != -1:
            raw = raw[start:end+1]
        result = json.loads(raw)
        return jsonify({"success": True, **result})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Follow-up Reminders ──────────────────────────────────────────────────────

@app.route("/api/reminders", methods=["GET"])
@login_required
def get_reminders():
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT r.*, l.name as lead_name, l.email as lead_email, l.country as lead_country
            FROM reminders r JOIN leads l ON r.lead_id = l.id
            ORDER BY r.remind_at ASC
        """)
        reminders = cursor.fetchall()
        for r in reminders:
            if r.get("remind_at"):
                r["remind_at"] = r["remind_at"].isoformat()
            if r.get("created_at"):
                r["created_at"] = r["created_at"].isoformat()
        cursor.close()
        conn.close()
        return jsonify({"success": True, "reminders": reminders})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/reminders", methods=["POST"])
@login_required
def add_reminder():
    try:
        data = request.get_json()
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO reminders (lead_id, note, remind_at, created_by)
            VALUES (%s, %s, %s, %s)
        """, (data.get("lead_id"), data.get("note"), data.get("remind_at"), data.get("created_by", "Team")))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/reminders/<int:reminder_id>", methods=["DELETE"])
@login_required
def delete_reminder(reminder_id):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM reminders WHERE id=%s", (reminder_id,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Competitor Tracker ───────────────────────────────────────────────────────

@app.route("/api/competitors", methods=["POST"])
@login_required
def track_competitors():
    try:
        data = request.get_json()
        niche = data.get("niche", "organic apparel")
        country = data.get("country", "India")
        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
        prompt = f"""Use web search to find real competitors selling {niche} in {country} and worldwide.
Search Google, Instagram, IndiaMART, Alibaba, Etsy, Amazon, Flipkart and social media.

Return ONLY a valid JSON array of competitors:
[
  {{
    "name": "Competitor brand name",
    "website": "their website",
    "instagram": "@handle",
    "country": "where they are based",
    "products": "what they sell",
    "price_range": "their pricing",
    "monthly_revenue": "estimated revenue if available",
    "strengths": "what they do well",
    "weaknesses": "where they are weak — your opportunity",
    "target_market": "who they sell to",
    "unique_angle": "their USP",
    "threat_level": "Low / Medium / High",
    "opportunity": "How Silasya/Shumitra can beat them or capture their customers"
  }}
]
Return ONLY the JSON array, no other text."""

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4000,
            tools=[{"type": "web_search_20250305", "name": "web_search"}],
            messages=[{"role": "user", "content": prompt}]
        )
        raw = ""
        for block in message.content:
            if hasattr(block, "text"):
                raw += block.text
        raw = raw.strip().replace("```json","").replace("```","").strip()
        start = raw.find("[")
        end = raw.rfind("]")
        if start != -1 and end != -1:
            raw = raw[start:end+1]
        competitors = json.loads(raw)
        return jsonify({"success": True, "competitors": competitors})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Export Excel ─────────────────────────────────────────────────────────────

@app.route("/api/export/excel")
@login_required
def export_excel():
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM leads ORDER BY created_at DESC")
        leads = cursor.fetchall()
        cursor.close()
        conn.close()

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Silasya Leads"
            headers = ["ID","Name","Type","Category","Country","City","Email","Phone","Website","Instagram","LinkedIn","Score","Status","Potential Value","Description","Why Good","Source","Saved By","Created"]
            gold = PatternFill("solid", fgColor="C9A84C")
            bold = Font(bold=True, color="000000")
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = gold
                cell.font = bold
                cell.alignment = Alignment(horizontal="center")
            for row, l in enumerate(leads, 2):
                ws.append([l.get("id"),l.get("name"),l.get("type"),l.get("category"),l.get("country"),l.get("city"),l.get("email"),l.get("phone"),l.get("website"),l.get("instagram"),l.get("linkedin"),l.get("score"),l.get("status"),l.get("potential_value"),l.get("description"),l.get("why_good"),l.get("source"),l.get("saved_by"),str(l.get("created_at",""))])
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 18
            import io
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return Response(buf.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                          headers={"Content-Disposition": "attachment;filename=silasya_leads.xlsx"})
        except ImportError:
            # Fallback to CSV if openpyxl not installed
            lines = ["Name,Type,Category,Country,City,Email,Phone,Score,Status"]
            for l in leads:
                lines.append(f'"{l.get("name","")}","{l.get("type","")}","{l.get("category","")}","{l.get("country","")}","{l.get("city","")}","{l.get("email","")}","{l.get("phone","")}","{l.get("score","")}","{l.get("status","")}"')
            return Response("\n".join(lines), mimetype="text/csv", headers={"Content-Disposition": "attachment;filename=silasya_leads.csv"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Update Lead Score ────────────────────────────────────────────────────────

@app.route("/api/leads/<int:lead_id>/score", methods=["PUT"])
@login_required
def update_score(lead_id):
    try:
        data = request.get_json()
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM leads WHERE id=%s", (lead_id,))
        lead = cursor.fetchone()
        if not lead:
            cursor.close()
            conn.close()
            return jsonify({"error": "Lead not found"}), 404

        # Auto-calculate score based on activity signals
        score = int(lead.get("score", 50))
        activity = data.get("activity", "")
        if activity == "opened_email": score = min(100, score + 10)
        elif activity == "replied": score = min(100, score + 25)
        elif activity == "clicked_link": score = min(100, score + 15)
        elif activity == "no_response_7days": score = max(0, score - 10)
        elif activity == "unsubscribed": score = max(0, score - 30)
        else:
            score = int(data.get("score", score))

        cursor2 = conn.cursor()
        cursor2.execute("UPDATE leads SET score=%s WHERE id=%s", (score, lead_id))
        conn.commit()
        cursor2.close()
        cursor.close()
        conn.close()
        return jsonify({"success": True, "new_score": score})
    except Exception as e:
        return jsonify({"error": str(e)}), 500




@app.errorhandler(400)
def bad_request(e):
    return jsonify({"error": "Bad request", "message": str(e)}), 400

@app.errorhandler(404)
def not_found(e):
    return jsonify({"error": "Endpoint not found", "message": str(e)}), 404

@app.errorhandler(405)
def method_not_allowed(e):
    return jsonify({"error": "Method not allowed", "message": str(e)}), 405

@app.errorhandler(500)
def internal_error(e):
    return jsonify({"error": "Internal server error", "message": str(e)}), 500

@app.errorhandler(Exception)
def handle_exception(e):
    return jsonify({"error": type(e).__name__, "message": str(e)}), 500


# ─── Main ────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("🚀 Starting Silasya & Shumitra Lead Finder...")
    print("📦 Initializing database...")
    init_db()
    port = int(os.getenv("PORT", 5000))
    debug = os.getenv("FLASK_DEBUG", "False").lower() == "true"  # Default OFF — debug=True bypasses error handlers
    print(f"✅ App running at http://localhost:{port}")
    print(f"📋 Share with team: http://YOUR_IP:{port}")
    app.run(host="0.0.0.0", port=port, debug=debug)