"""
Silasya & Shumitra — AI Lead Finder
Flask + MySQL Backend
"""

from flask import Flask, request, jsonify, render_template, session, Response
from flask_cors import CORS
import mysql.connector
from mysql.connector import pooling
import os
import json
import anthropic
from datetime import datetime
from dotenv import load_dotenv
import threading
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta

load_dotenv()

# ─── Simple In-Memory Cache ───────────────────────────────────────────────────
import hashlib, time
_cache = {}
CACHE_TTL = 86400  # 24 hours

def cache_get(key):
    if key in _cache:
        data, ts = _cache[key]
        if time.time() - ts < CACHE_TTL:
            return data
        del _cache[key]
    return None

def cache_set(key, data):
    _cache[key] = (data, time.time())

def make_key(*args):
    return hashlib.md5("|".join(str(a).lower().strip() for a in args).encode()).hexdigest()

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "silasya-secret-2025")
CORS(app)

# ─── DB Pool ─────────────────────────────────────────────────────────────────

db_config = {
    "host": os.getenv("DB_HOST", "localhost"),
    "port": int(os.getenv("DB_PORT", 3306)),
    "user": os.getenv("DB_USER", "root"),
    "password": os.getenv("DB_PASSWORD", ""),
    "database": os.getenv("DB_NAME", "silasya_leads"),
}

connection_pool = None


def get_pool():
    global connection_pool
    if connection_pool is None:
        connection_pool = pooling.MySQLConnectionPool(
            pool_name="silasya_pool",
            pool_size=5,
            **db_config
        )
    return connection_pool


def get_db():
    return get_pool().get_connection()


# ─── DB Init ─────────────────────────────────────────────────────────────────

def init_db():
    conn = mysql.connector.connect(
        host=db_config["host"],
        port=db_config["port"],
        user=db_config["user"],
        password=db_config["password"],
    )
    cursor = conn.cursor()
    cursor.execute(f"CREATE DATABASE IF NOT EXISTS `{db_config['database']}` CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci")
    cursor.execute(f"USE `{db_config['database']}`")
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS leads (
            id INT AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(255) NOT NULL,
            type VARCHAR(10) DEFAULT 'b2c',
            category VARCHAR(255),
            country VARCHAR(100),
            city VARCHAR(100),
            email VARCHAR(255),
            phone VARCHAR(100),
            website VARCHAR(500),
            instagram VARCHAR(255),
            linkedin VARCHAR(500),
            facebook VARCHAR(500),
            whatsapp VARCHAR(100),
            description TEXT,
            why_good TEXT,
            potential_value VARCHAR(100),
            tags JSON,
            score INT DEFAULT 50,
            status VARCHAR(50) DEFAULT 'new',
            source VARCHAR(255),
            notes TEXT,
            saved_by VARCHAR(100),
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS search_history (
            id INT AUTO_INCREMENT PRIMARY KEY,
            business VARCHAR(50),
            country VARCHAR(255),
            niche VARCHAR(255),
            lead_type VARCHAR(255),
            channels TEXT,
            keywords TEXT,
            leads_found INT DEFAULT 0,
            searched_by VARCHAR(100),
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS team_members (
            id INT AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(100) NOT NULL,
            email VARCHAR(255) UNIQUE NOT NULL,
            role VARCHAR(20) DEFAULT 'member',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)
    cursor.execute("SELECT COUNT(*) FROM team_members")
    if cursor.fetchone()[0] == 0:
        cursor.execute("INSERT INTO team_members (name, email, role) VALUES ('Admin', 'admin@silasya.com', 'admin')")
    conn.commit()
    cursor.close()
    conn.close()
    print("✅ Database initialized successfully")


# ─── Routes ──────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/favicon.ico")
def favicon():
    return app.send_static_file("favicon.ico")

@app.route("/manifest.json")
def manifest():
    return app.send_static_file("manifest.json")

@app.route("/sw.js")
def sw():
    return app.send_static_file("sw.js")


@app.route("/api/status")
def status():
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM leads")
        total = cursor.fetchone()[0]
        cursor.close()
        conn.close()
        return jsonify({"status": "ok", "total_leads": total, "db": "connected"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


# ─── AI Search ───────────────────────────────────────────────────────────────

@app.route("/api/search", methods=["POST"])
def ai_search():
    try:
        data = request.get_json()
        business = data.get("business", "both")
        country = data.get("country", "India and worldwide")
        niche = data.get("niche", "organic products")
        lead_type = data.get("lead_type", "all")
        channels = data.get("channels", "Google, LinkedIn, Instagram")
        keywords = data.get("keywords", "")
        saved_by = data.get("saved_by", "Team")

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO search_history (business, country, niche, lead_type, channels, keywords, searched_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (business, country, niche, lead_type, channels, keywords, saved_by))
        conn.commit()
        search_id = cursor.lastrowid

        cursor.execute("""
            DELETE FROM search_history WHERE id NOT IN (
                SELECT id FROM (SELECT id FROM search_history ORDER BY created_at DESC LIMIT 100) AS tmp
            )
        """)
        conn.commit()
        cursor.close()
        conn.close()

        # Check cache first
        cache_key = make_key("leads", business, country, niche, lead_type, keywords)
        cached = cache_get(cache_key)
        if cached:
            return jsonify({"success": True, "leads": cached, "count": len(cached), "cached": True})

        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

        prompt = f"""You are a lead generation expert for two Indian businesses:
- SILASYA: B2C organic apparel, toys, home decor brand (baby products, kids toys, home decor, organic clothing) — focuses on organic/eco products only
- SHOUMITRA: B2B export and supply arm — fulfills ANY demand, organic or conventional. Exports vegetables, fruits, spices, grains, medicines, Ayurvedic products, textiles, apparel, home goods, handicrafts, chemicals, raw materials — whatever the buyer needs. Shoumitra sources and supplies it. No restriction on product type.

Generate 15 realistic, detailed leads based on these parameters:
Business Focus: {business}
Target Country/Region: {country}
Product Niche: {niche}
Lead Type: {lead_type}
Search Channels: {channels}
Extra Keywords: {keywords}

Search across ALL these sources to find leads:
MARKETPLACES: Amazon, Flipkart, Meesho, Etsy, IndiaMART, Alibaba, TradeIndia, Global Sources, EC21, ExportHub, Faire, Handshake
SOCIAL MEDIA: Instagram shops, Facebook marketplace, LinkedIn company pages, Pinterest shops
GOVERNMENT: DGFT exporters list, Startup India, MSME registry, GeM portal buyers, export promotion councils (EPCH, AEPC, APEDA)
CORPORATE GIFTING: Corporate gifting companies, HR departments, event management firms, wedding planners, festival gifting buyers
RETAIL: Organic stores, eco stores, boutiques, department stores, supermarket chains (Nature's Basket, Godrej Nature's Basket, BigBasket organic)
HOSPITALITY: Hotels, resorts, spas, wellness centers, yoga studios looking for organic products
SCHOOLS & INSTITUTIONS: Schools, NGOs, hospitals looking for organic/eco products for kids
INTERNATIONAL: Import companies, distributors, wholesalers, fair trade organizations

Return ONLY a valid JSON array with exactly 15 leads. Each lead must have:
- name (string): Company or person name
- type (string): "b2c" or "b2b"
- category (string): be specific e.g. "Amazon Seller", "Corporate Gifting", "Government Buyer", "Hotel Chain", "Organic Supermarket", "Export Distributor", "NGO", "Wedding Planner"
- country (string)
- city (string)
- email (string): realistic business email
- phone (string): realistic phone with country code
- website (string): realistic URL
- instagram (string): @handle
- linkedin (string): realistic LinkedIn URL
- facebook (string): realistic Facebook page URL
- whatsapp (string): phone number
- description (string): 1-2 sentence description
- why_good (string): why this is a good lead for Silasya/Shoumitra specifically
- potential_value (string): estimated monthly value e.g. "$5,000-$10,000/month"
- score (number): 1-100 lead quality score
- tags (array of strings): relevant tags including source platform
- source (string): exact platform/website where this lead was found

Return ONLY the JSON array, no other text."""

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=6000,
            messages=[{"role": "user", "content": prompt}]
        )

        raw = message.content[0].text.strip()
        raw = raw.replace("```json", "").replace("```", "").strip()
        start = raw.find("[")
        end = raw.rfind("]")
        if start != -1 and end != -1:
            raw = raw[start:end+1]

        leads = json.loads(raw)

        # Update search history with count
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("UPDATE search_history SET leads_found = %s WHERE id = %s", (len(leads), search_id))
        conn.commit()
        cursor.close()
        conn.close()

        cache_set(cache_key, leads)
        return jsonify({"success": True, "leads": leads, "count": len(leads)})

    except json.JSONDecodeError as e:
        return jsonify({"error": f"AI returned invalid data: {str(e)}"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Save Lead ───────────────────────────────────────────────────────────────

@app.route("/api/leads", methods=["GET"])
def get_leads():
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        status_filter = request.args.get("status", "")
        type_filter = request.args.get("type", "")
        search = request.args.get("search", "")

        query = "SELECT * FROM leads WHERE 1=1"
        params = []

        if status_filter:
            query += " AND status = %s"
            params.append(status_filter)
        if type_filter:
            query += " AND type = %s"
            params.append(type_filter)
        if search:
            query += " AND (name LIKE %s OR email LIKE %s OR country LIKE %s OR category LIKE %s)"
            s = f"%{search}%"
            params.extend([s, s, s, s])
        category_filter = request.args.get("category", "")
        if category_filter:
            query += " AND category = %s"
            params.append(category_filter)

        query += " ORDER BY created_at DESC"
        cursor.execute(query, params)
        leads = cursor.fetchall()

        for lead in leads:
            if lead.get("tags") and isinstance(lead["tags"], str):
                try:
                    lead["tags"] = json.loads(lead["tags"])
                except:
                    lead["tags"] = []
            if lead.get("created_at"):
                lead["created_at"] = lead["created_at"].isoformat()
            if lead.get("updated_at"):
                lead["updated_at"] = lead["updated_at"].isoformat()

        cursor.close()
        conn.close()
        return jsonify({"success": True, "leads": leads, "count": len(leads)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/leads", methods=["POST"])
def save_lead():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No data provided"}), 400

        name = data.get("name") or "Unknown Lead"
        email = data.get("email", "")
        phone = data.get("phone", "")
        country = data.get("country", "")
        tags = data.get("tags", [])
        if isinstance(tags, list):
            tags = json.dumps(tags)

        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        # Check for duplicate by email
        if email:
            cursor.execute("SELECT id, name FROM leads WHERE email = %s", (email,))
            existing = cursor.fetchone()
            if existing:
                cursor.close()
                conn.close()
                return jsonify({
                    "success": False,
                    "duplicate": True,
                    "id": existing["id"],
                    "message": f"'{name}' already saved (duplicate email)"
                })

        # Check for duplicate by name + country
        cursor.execute("SELECT id, name FROM leads WHERE name = %s AND country = %s", (name, country))
        existing = cursor.fetchone()
        if existing:
            cursor.close()
            conn.close()
            return jsonify({
                "success": False,
                "duplicate": True,
                "id": existing["id"],
                "message": f"'{name}' from {country} already exists in Lead Bank"
            })

        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO leads (name, type, category, country, city, email, phone, website,
                instagram, linkedin, facebook, whatsapp, description, why_good,
                potential_value, tags, score, status, source, saved_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            name,
            data.get("type", "b2c"),
            data.get("category", ""),
            country,
            data.get("city", ""),
            email,
            phone,
            data.get("website", ""),
            data.get("instagram", ""),
            data.get("linkedin", ""),
            data.get("facebook", ""),
            data.get("whatsapp", ""),
            data.get("description", ""),
            data.get("why_good", ""),
            data.get("potential_value", ""),
            tags,
            int(data.get("score", 50)),
            data.get("status", "new"),
            data.get("source", "AI Search"),
            data.get("saved_by", "Team"),
        ))
        conn.commit()
        lead_id = cursor.lastrowid
        cursor.close()
        conn.close()

        return jsonify({"success": True, "id": lead_id, "message": f"Lead '{name}' saved!"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/leads/<int:lead_id>", methods=["PUT"])
def update_lead(lead_id):
    try:
        data = request.get_json()
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE leads SET status=%s, notes=%s WHERE id=%s
        """, (data.get("status", "new"), data.get("notes", ""), lead_id))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/leads/<int:lead_id>", methods=["DELETE"])
def delete_lead(lead_id):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM leads WHERE id=%s", (lead_id,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Stats ───────────────────────────────────────────────────────────────────

@app.route("/api/stats")
def get_stats():
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT COUNT(*) as total FROM leads")
        total = cursor.fetchone()["total"]
        cursor.execute("SELECT COUNT(*) as hot FROM leads WHERE score >= 80")
        hot = cursor.fetchone()["hot"]
        cursor.execute("SELECT COUNT(*) as converted FROM leads WHERE status = 'converted'")
        converted = cursor.fetchone()["converted"]
        cursor.execute("SELECT COUNT(DISTINCT country) as countries FROM leads")
        countries = cursor.fetchone()["countries"]
        cursor.close()
        conn.close()
        return jsonify({"total": total, "hot": hot, "converted": converted, "countries": countries})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Export CSV ──────────────────────────────────────────────────────────────

@app.route("/api/export/csv")
def export_csv():
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM leads ORDER BY created_at DESC")
        leads = cursor.fetchall()
        cursor.close()
        conn.close()

        lines = ["Name,Type,Category,Country,City,Email,Phone,Website,Score,Status,Saved By,Created At"]
        for l in leads:
            lines.append(f'"{l.get("name","")}", "{l.get("type","")}", "{l.get("category","")}", "{l.get("country","")}", "{l.get("city","")}", "{l.get("email","")}", "{l.get("phone","")}", "{l.get("website","")}", "{l.get("score","")}", "{l.get("status","")}", "{l.get("saved_by","")}", "{l.get("created_at","")}"')

        csv = "\n".join(lines)
        return Response(csv, mimetype="text/csv", headers={"Content-Disposition": "attachment;filename=silasya_leads.csv"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Team ────────────────────────────────────────────────────────────────────

@app.route("/api/team", methods=["GET"])
def get_team():
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM team_members ORDER BY created_at")
        members = cursor.fetchall()
        for m in members:
            if m.get("created_at"):
                m["created_at"] = m["created_at"].isoformat()
        cursor.close()
        conn.close()
        return jsonify({"success": True, "members": members})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/team", methods=["POST"])
def add_team():
    try:
        data = request.get_json()
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO team_members (name, email, role) VALUES (%s,%s,%s)
        """, (data.get("name"), data.get("email"), data.get("role", "member")))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── AI Outreach ─────────────────────────────────────────────────────────────

@app.route("/api/outreach/<int:lead_id>", methods=["POST"])
def generate_outreach(lead_id):
    try:
        data = request.get_json()
        channel = data.get("channel", "email")

        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM leads WHERE id=%s", (lead_id,))
        lead = cursor.fetchone()
        cursor.close()
        conn.close()

        if not lead:
            return jsonify({"error": "Lead not found"}), 404

        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

        prompt = f"""Write a {channel} outreach message for this lead on behalf of Silasya & Shoumitra (Indian organic brand).

Lead: {lead.get('name')}
Type: {lead.get('type')}
Category: {lead.get('category')}
Country: {lead.get('country')}
Description: {lead.get('description')}

Write a short, friendly, professional {channel} message. Keep it under 150 words."""

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=500,
            messages=[{"role": "user", "content": prompt}]
        )

        return jsonify({"success": True, "message": message.content[0].text})
    except Exception as e:
        return jsonify({"error": str(e)}), 500



# ─── Bulk Save Leads ─────────────────────────────────────────────────────────

@app.route("/api/leads/bulk", methods=["POST"])
def save_leads_bulk():
    try:
        data = request.get_json()
        leads = data.get("leads", [])
        saved_by = data.get("saved_by", "Team")
        if not leads:
            return jsonify({"error": "No leads provided"}), 400
        conn = get_db()
        cursor = conn.cursor()
        saved_ids = []
        for lead in leads:
            tags = lead.get("tags", [])
            if isinstance(tags, list):
                tags = json.dumps(tags)
            cursor.execute("""
                INSERT INTO leads (name, type, category, country, city, email, phone, website,
                    instagram, linkedin, facebook, whatsapp, description, why_good,
                    potential_value, tags, score, status, source, saved_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                lead.get("name") or "Unknown Lead",
                lead.get("type", "b2c"), lead.get("category", ""),
                lead.get("country", ""), lead.get("city", ""),
                lead.get("email", ""), lead.get("phone", ""),
                lead.get("website", ""), lead.get("instagram", ""),
                lead.get("linkedin", ""), lead.get("facebook", ""),
                lead.get("whatsapp", ""), lead.get("description", ""),
                lead.get("why_good", ""), lead.get("potential_value", ""),
                tags, int(lead.get("score", 50)),
                lead.get("status", "new"), lead.get("source", "AI Search"), saved_by,
            ))
            saved_ids.append(cursor.lastrowid)
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True, "saved": len(saved_ids), "ids": saved_ids})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Export Excel ─────────────────────────────────────────────────────────────

@app.route("/api/export/excel")
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import io
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM leads ORDER BY created_at DESC")
        leads = cursor.fetchall()
        cursor.close()
        conn.close()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Silasya Leads"
        headers = ["ID","Name","Type","Category","Country","City","Email","Phone",
                   "Website","Instagram","LinkedIn","WhatsApp","Score","Status",
                   "Potential Value","Description","Why Good","Saved By","Created At"]
        ws.append(headers)
        gold = PatternFill("solid", fgColor="C9A84C")
        bold = Font(bold=True, color="000000")
        for cell in ws[1]:
            cell.fill = gold
            cell.font = bold
            cell.alignment = Alignment(horizontal="center")
        for lead in leads:
            ws.append([
                lead.get("id"), lead.get("name"), lead.get("type"),
                lead.get("category"), lead.get("country"), lead.get("city"),
                lead.get("email"), lead.get("phone"), lead.get("website"),
                lead.get("instagram"), lead.get("linkedin"), lead.get("whatsapp"),
                lead.get("score"), lead.get("status"), lead.get("potential_value"),
                lead.get("description"), lead.get("why_good"), lead.get("saved_by"),
                str(lead.get("created_at", ""))
            ])
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(buf.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment;filename=silasya_leads.xlsx"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Email Outreach ───────────────────────────────────────────────────────────

@app.route("/api/email/<int:lead_id>", methods=["POST"])
def generate_email(lead_id):
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM leads WHERE id=%s", (lead_id,))
        lead = cursor.fetchone()
        cursor.close()
        conn.close()
        if not lead:
            return jsonify({"error": "Lead not found"}), 404
        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
        prompt = f"""Write a professional cold email for Silasya & Shoumitra (Indian organic brand).
Lead: {lead.get('name')}, {lead.get('category')}, {lead.get('country')}
Description: {lead.get('description')}
Write subject line + email body. Under 150 words. Friendly and focused on partnership."""
        message = client.messages.create(
            model="claude-sonnet-4-20250514", max_tokens=500,
            messages=[{"role": "user", "content": prompt}])
        return jsonify({"success": True, "message": message.content[0].text})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Reminders ───────────────────────────────────────────────────────────────

@app.route("/api/reminders", methods=["GET"])
def get_reminders():
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT r.*, l.name as lead_name FROM reminders r
            LEFT JOIN leads l ON r.lead_id = l.id
            ORDER BY r.remind_at ASC
        """)
        reminders = cursor.fetchall()
        for r in reminders:
            if r.get("remind_at"): r["remind_at"] = r["remind_at"].isoformat()
            if r.get("created_at"): r["created_at"] = r["created_at"].isoformat()
        cursor.close()
        conn.close()
        return jsonify({"success": True, "reminders": reminders})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/reminders", methods=["POST"])
def add_reminder():
    try:
        data = request.get_json()
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO reminders (lead_id, note, remind_at, created_by)
            VALUES (%s, %s, %s, %s)
        """, (data.get("lead_id"), data.get("note",""), data.get("remind_at"), data.get("created_by","Team")))
        conn.commit()
        rid = cursor.lastrowid
        cursor.close()
        conn.close()
        return jsonify({"success": True, "id": rid})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/reminders/<int:reminder_id>", methods=["DELETE"])
def delete_reminder(reminder_id):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM reminders WHERE id=%s", (reminder_id,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/reminders/<int:reminder_id>", methods=["PUT"])
def update_reminder(reminder_id):
    try:
        data = request.get_json()
        conn = get_db()
        cursor = conn.cursor()
        if "remind_at" in data:
            cursor.execute("UPDATE reminders SET done=%s, remind_at=%s WHERE id=%s",
                         (data.get("done", 0), data.get("remind_at"), reminder_id))
        else:
            cursor.execute("UPDATE reminders SET done=%s WHERE id=%s",
                         (data.get("done", 1), reminder_id))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Login ────────────────────────────────────────────────────────────────────

TEAM_CREDENTIALS = {
    os.getenv("LOGIN_USER_1", "silasya"):   os.getenv("LOGIN_PASS_1", "silasya2025"),
    os.getenv("LOGIN_USER_2", "shoumitra"): os.getenv("LOGIN_PASS_2", "shoumitra2025"),
    os.getenv("LOGIN_USER_3", "admin"):     os.getenv("LOGIN_PASS_3", "admin2025"),
}

@app.route("/login", methods=["GET"])
def login_page():
    return render_template("login.html")

@app.route("/login", methods=["POST"])
def do_login():
    data = request.get_json()
    username = (data.get("username") or "").strip().lower()
    password = data.get("password") or ""
    if TEAM_CREDENTIALS.get(username) == password:
        session["user"] = username
        return jsonify({"success": True})
    return jsonify({"success": False}), 401

@app.route("/logout")
def logout():
    session.clear()
    return jsonify({"success": True})



# ─── Buyer Requirements ───────────────────────────────────────────────────────



@app.route("/api/team/<int:member_id>", methods=["DELETE"])
def delete_team_member(member_id):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM team_members WHERE id=%s AND email != 'admin@silasya.com'", (member_id,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/buyer-requirements", methods=["POST"])
def buyer_requirements():
    try:
        data = request.get_json()
        niche = data.get("niche", "organic products")
        country = data.get("country", "worldwide")
        business = data.get("business", "both")
        buyer_type = data.get("buyer_type", "all")
        keywords = data.get("keywords", "")

        # Check cache first
        cache_key = make_key("rfq", niche, country, business, buyer_type, keywords)
        cached = cache_get(cache_key)
        if cached:
            return jsonify({"success": True, "results": cached, "count": len(cached), "cached": True})

        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
        prompt = f"""You are a B2B sourcing expert for Shoumitra — an Indian export and supply company that fulfills ANY buyer demand (organic or conventional). Shoumitra can supply: fresh vegetables, fruits, spices, grains, pulses, medicines, Ayurvedic products, organic products, textiles, apparel, home decor, handicrafts, chemicals, raw materials, processed foods, dry fruits — anything a buyer needs. Find 10 realistic buyer requirements for:
Business Focus: {business}
Product Niche: {niche}
Target Country: {country}
Preferred Buyer Type: {buyer_type}
Extra Keywords: {keywords}

Search across ALL these buyer sources:
TRADE PORTALS: IndiaMART, Alibaba, TradeIndia, Global Sources, EC21, ExportHub, Faire, Handshake
GOVERNMENT: GeM portal bulk buyers, government school/hospital tenders, DGFT registered importers, CSR bulk buyers, export promotion councils (EPCH, AEPC, APEDA)
CORPORATE GIFTING: Corporate gifting companies, HR departments buying festival/Diwali gifts, event management firms, wedding planners buying bulk gifts
HOSPITALITY: Hotels, resorts, spas, wellness centers, yoga studios, Ayurveda centers buying organic products in bulk
RETAIL CHAINS: Supermarket chains, organic store chains (Nature's Basket, Whole Foods), department stores, boutique chains
SCHOOLS & INSTITUTIONS: Schools, NGOs, hospitals, charitable trusts, orphanages buying organic/eco products
INTERNATIONAL: Fair trade organizations, organic importers, wholesale distributors, Amazon FBA sellers

Return ONLY a valid JSON array with 10 items. Each item must have:
- buyer_name (string): buyer company or person name
- buyer_type (string): e.g. "Corporate Gifting", "Government Tender", "Hotel Chain", "NGO", "School", "Retail Chain", "International Importer", "Wedding Planner", "Spa & Wellness"
- country (string)
- city (string)
- requirement (string): exactly what they need
- quantity (string): e.g. "500 units/month"
- budget (string): e.g. "$5,000-$10,000"
- contact_email (string): realistic business email
- phone (string): realistic phone with country code
- whatsapp (string): same as phone
- website (string): realistic company website URL
- linkedin (string): realistic LinkedIn URL e.g. "https://linkedin.com/company/name"
- facebook (string): realistic Facebook page URL
- platform (string): exact platform/source where this RFQ was found
- urgency (string): "High", "Medium", or "Low"
- posted (string): e.g. "2 days ago"
- match_score (number): 0-100 relevance score
- notes (string): certifications needed, special requirements, additional details

Return ONLY the JSON array, no other text."""

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=6000,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = message.content[0].text.strip()
        raw = raw.replace("```json", "").replace("```", "").strip()
        start = raw.find("[")
        end = raw.rfind("]")
        if start != -1 and end != -1:
            raw = raw[start:end+1]
        results = json.loads(raw)
        cache_set(cache_key, results)
        return jsonify({"success": True, "results": results, "count": len(results)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Competitor Tracker ───────────────────────────────────────────────────────

@app.route("/api/competitors", methods=["POST"])
def find_competitors():
    try:
        data = request.get_json()
        niche = data.get("niche", "organic products")
        country = data.get("country", "India")
        business = data.get("business", "both")
        comp_type = data.get("comp_type", "all")
        keywords = data.get("keywords", "")

        # Check cache first
        cache_key = make_key("comp", niche, country, business, comp_type, keywords)
        cached = cache_get(cache_key)
        if cached:
            return jsonify({"success": True, "results": cached, "count": len(cached), "cached": True})

        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
        prompt = f"""You are a competitive intelligence expert. Find 10 competitors for:
- SILASYA: Indian organic B2C brand (apparel, toys, home decor)
- SHOUMITRA: Indian B2B export/supply company that fulfills ANY demand — vegetables, fruits, spices, medicines, textiles, handicrafts, chemicals, raw materials, processed foods — organic or conventional

Find competitors in:
Business Focus: {business}
Product Niche: {niche}
Country: {country}
Competitor Type Focus: {comp_type}
Extra Keywords: {keywords}

Search across ALL these channels:
MARKETPLACES: Amazon sellers, Flipkart sellers, Meesho sellers, Etsy shops, Faire brands
SOCIAL MEDIA: Instagram organic brands, Facebook shops, Pinterest sellers
D2C WEBSITES: Brands selling directly on their own website
EXPORT PLATFORMS: Alibaba stores, IndiaMART suppliers, TradeIndia sellers
GOVERNMENT REGISTERED: MSME registered organic brands, Startup India registered, GI tagged products
RETAIL: Brands in Nature's Basket, Whole Foods, organic stores, boutiques
HOSPITALITY SUPPLIERS: Brands supplying hotels, spas, wellness centers
CORPORATE GIFTING BRANDS: Brands focused on corporate gifting market
INTERNATIONAL: Fair trade certified brands, organic certified exporters

Return ONLY a valid JSON array with 10 items. Each item must have:
- name (string): competitor brand name
- country (string)
- city (string): their main city
- website (string): realistic URL starting with https://
- instagram (string): @handle without spaces
- linkedin (string): realistic LinkedIn URL
- facebook (string): realistic Facebook page URL
- email (string): realistic contact email
- phone (string): realistic phone with country code
- whatsapp (string): same as phone
- price_range (string): e.g. "₹500-₹2000" or "$10-$50"
- products (string): exactly what they sell
- strengths (string): their main strength in 1 sentence
- weaknesses (string): their main weakness in 1 sentence
- how_to_beat (string): specific actionable strategy to beat them
- threat_level (string): "High", "Medium", or "Low"
- platform (string): their strongest sales channel
- selling_channels (string): all platforms they sell on
- monthly_revenue (string): estimated monthly revenue
- target_market (string): exactly who they sell to
- certifications (string): any certifications e.g. "GOTS, OEKO-TEX, Fair Trade"

Return ONLY the JSON array, no other text."""

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=6000,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = message.content[0].text.strip()
        raw = raw.replace("```json", "").replace("```", "").strip()
        start = raw.find("[")
        end = raw.rfind("]")
        if start != -1 and end != -1:
            raw = raw[start:end+1]
        results = json.loads(raw)
        cache_set(cache_key, results)
        return jsonify({"success": True, "results": results, "count": len(results)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500






# ─── Email Notifications ──────────────────────────────────────────────────────

def get_team_emails():
    """Get all team member emails from DB"""
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT email, name FROM team_members WHERE email != 'admin@silasya.com'")
        members = cursor.fetchall()
        cursor.close()
        conn.close()
        return members
    except:
        return []

def send_email_notification(subject, html_body):
    """Send email via Resend API over HTTPS"""
    resend_key = os.getenv("RESEND_API_KEY", "").strip()
    if not resend_key:
        print("RESEND_API_KEY not set")
        return False
    members = get_team_emails()
    if not members:
        print("No team members")
        return False
    sender = os.getenv("MAIL_SENDER", "onboarding@resend.dev").strip()
    success_count = 0
    for member in members:
        try:
            personal_html = html_body.replace("{{name}}", member["name"])
            r = requests.post(
                "https://api.resend.com/emails",
                headers={"Authorization": f"Bearer {resend_key}", "Content-Type": "application/json"},
                json={"from": f"Silasya Lead Finder <{sender}>", "to": [member["email"]], "subject": subject, "html": personal_html},
                timeout=10
            )
            if r.status_code in [200, 201]:
                print(f"Email sent to {member['email']}")
                success_count += 1
            else:
                print(f"Email failed to {member['email']}: {r.text}")
        except Exception as e:
            print(f"Email error: {e}")
    return success_count > 0

def send_new_leads_email(leads, niche, country):
    """Send email when new leads are found"""
    if not leads:
        return
    
    leads_html = "".join([f"""
    <tr>
      <td style="padding:10px;border-bottom:1px solid #eee">
        <strong style="color:#1a1a1a">{l.get('name','')}</strong><br>
        <span style="color:#666;font-size:12px">{l.get('category','')} · {l.get('country','')}</span>
      </td>
      <td style="padding:10px;border-bottom:1px solid #eee;text-align:center">
        <span style="background:{'#cf4c4c' if int(l.get('score',0))>=80 else '#c9a84c' if int(l.get('score',0))>=60 else '#888'};color:white;padding:3px 8px;border-radius:4px;font-size:12px;font-weight:700">{l.get('score','')}</span>
      </td>
      <td style="padding:10px;border-bottom:1px solid #eee;font-size:12px;color:#666">{l.get('potential_value','')}</td>
      <td style="padding:10px;border-bottom:1px solid #eee;font-size:12px;color:#666">{l.get('source','')}</td>
    </tr>
    """ for l in leads[:10]])

    html = f"""
    <!DOCTYPE html>
    <html>
    <body style="font-family:Arial,sans-serif;background:#f5f5f5;padding:20px">
      <div style="max-width:600px;margin:0 auto;background:white;border-radius:12px;overflow:hidden">
        <div style="background:#0a0a0b;padding:24px;text-align:center">
          <h1 style="color:#c9a84c;margin:0;font-size:22px">SILASYA × SHOUMITRA</h1>
          <p style="color:#6e6b67;margin:4px 0 0;font-size:12px">AI Lead Finder</p>
        </div>
        <div style="padding:24px">
          <h2 style="color:#1a1a1a;margin:0 0 8px">🤖 {len(leads)} New Leads Found!</h2>
          <p style="color:#666;margin:0 0 20px">Auto-search found new leads for <strong>{niche}</strong> in <strong>{country}</strong></p>
          <p style="color:#888;font-size:13px;margin:0 0 16px">Hi {{{{name}}}}, here are the latest leads:</p>
          <table style="width:100%;border-collapse:collapse;font-size:13px">
            <thead>
              <tr style="background:#f9f9f9">
                <th style="padding:10px;text-align:left;color:#666;font-weight:600">Lead</th>
                <th style="padding:10px;text-align:center;color:#666;font-weight:600">Score</th>
                <th style="padding:10px;text-align:left;color:#666;font-weight:600">Value</th>
                <th style="padding:10px;text-align:left;color:#666;font-weight:600">Source</th>
              </tr>
            </thead>
            <tbody>{leads_html}</tbody>
          </table>
          <div style="margin-top:24px;text-align:center">
            <a href="{os.getenv('APP_URL','https://lead-generator-production-5218.up.railway.app')}" 
               style="background:#c9a84c;color:#0a0a0b;padding:12px 28px;border-radius:8px;text-decoration:none;font-weight:700;font-size:14px">
              View All Leads →
            </a>
          </div>
        </div>
        <div style="background:#f9f9f9;padding:16px;text-align:center;font-size:11px;color:#999">
          You received this because you are a team member of Silasya × Shoumitra Lead Finder.<br>
          Contact your admin to unsubscribe.
        </div>
      </div>
    </body>
    </html>
    """
    
    hot_count = len([l for l in leads if int(l.get('score',0)) >= 80])
    subject = f"🔥 {len(leads)} New Leads Found ({hot_count} Hot!) — {niche}"
    send_email_notification(subject, html)

# ─── Notifications ────────────────────────────────────────────────────────────

@app.route("/api/notifications", methods=["GET"])
def get_notifications():
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        # Create table if not exists
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS notifications (
                id INT AUTO_INCREMENT PRIMARY KEY,
                title VARCHAR(255),
                message TEXT,
                type VARCHAR(50) DEFAULT 'info',
                is_read TINYINT DEFAULT 0,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()
        cursor.execute("SELECT * FROM notifications ORDER BY created_at DESC LIMIT 20")
        notifs = cursor.fetchall()
        cursor.execute("SELECT COUNT(*) as unread FROM notifications WHERE is_read=0")
        unread = cursor.fetchone()["unread"]
        for n in notifs:
            if n.get("created_at"):
                n["created_at"] = n["created_at"].isoformat()
        cursor.close()
        conn.close()
        return jsonify({"success": True, "notifications": notifs, "unread": unread})
    except Exception as e:
        return jsonify({"success": True, "notifications": [], "unread": 0})


@app.route("/api/notifications/read", methods=["POST"])
def mark_notifications_read():
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("UPDATE notifications SET is_read=1")
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def add_notification(title, message, notif_type="info"):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO notifications (title, message, type) VALUES (%s,%s,%s)",
            (title, message, notif_type)
        )
        # Keep only last 50 notifications
        cursor.execute("""
            DELETE FROM notifications WHERE id NOT IN (
                SELECT id FROM (SELECT id FROM notifications ORDER BY created_at DESC LIMIT 50) AS tmp
            )
        """)
        conn.commit()
        cursor.close()
        conn.close()
    except Exception as e:
        print(f"Notification error: {e}")


# ─── Auto Search ──────────────────────────────────────────────────────────────

AUTO_SEARCH_NICHES = [
    "Organic Apparel / Clothing",
    "Home Décor / Sustainable Living",
    "Organic Toys / Kids",
    "Organic Gift Sets / Lifestyle",
    "Fresh Vegetables / Fruits",
    "Spices / Condiments",
]

def run_auto_search():
    """Runs every 6 hours — finds new leads and saves them automatically"""
    while True:
        try:
            # Wait 6 hours
            threading.Event().wait(6 * 60 * 60)

            print("🤖 Auto-search running...")

            api_key = os.getenv("ANTHROPIC_API_KEY", "").strip()
            if not api_key or not api_key.startswith("sk-ant-"):
                print("❌ Auto-search skipped: No API key")
                continue

            # Pick a niche to search (rotate through them)
            import random
            niche = random.choice(AUTO_SEARCH_NICHES)
            countries = ["India", "Germany", "USA", "UAE", "UK"]
            country = random.choice(countries)

            client = anthropic.Anthropic(api_key=api_key)
            prompt = f"""You are a lead generation expert for Silasya (organic B2C) and Shoumitra (B2B export).
Find 5 fresh, realistic leads for:
Niche: {niche}
Country: {country}

Return ONLY a valid JSON array with 5 leads. Each must have:
- name, type (b2c/b2b), category, country, city, email, phone, website, description, why_good, potential_value, score (1-100), tags (array), source
Return ONLY the JSON array."""

            message = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=3000,
                messages=[{"role": "user", "content": prompt}]
            )

            raw = message.content[0].text.strip()
            raw = raw.replace("```json", "").replace("```", "").strip()
            start = raw.find("[")
            end = raw.rfind("]")
            if start != -1 and end != -1:
                raw = raw[start:end+1]

            leads = json.loads(raw)
            saved_count = 0

            conn = get_db()
            cursor = conn.cursor(dictionary=True)

            for lead in leads:
                name = lead.get("name") or "Unknown"
                email = lead.get("email", "")
                country_val = lead.get("country", "")

                # Check duplicate
                if email:
                    cursor.execute("SELECT id FROM leads WHERE email=%s", (email,))
                    if cursor.fetchone():
                        continue
                cursor.execute("SELECT id FROM leads WHERE name=%s AND country=%s", (name, country_val))
                if cursor.fetchone():
                    continue

                tags = lead.get("tags", [])
                if isinstance(tags, list):
                    tags = json.dumps(tags)

                cursor.execute("""
                    INSERT INTO leads (name, type, category, country, city, email, phone,
                        website, description, why_good, potential_value, tags, score,
                        status, source, saved_by)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    name, lead.get("type","b2b"), lead.get("category",""),
                    country_val, lead.get("city",""), email,
                    lead.get("phone",""), lead.get("website",""),
                    lead.get("description",""), lead.get("why_good",""),
                    lead.get("potential_value",""), tags,
                    int(lead.get("score",50)), "new",
                    lead.get("source","Auto Search"), "Auto Search"
                ))
                saved_count += 1

            conn.commit()

            # Log the search
            cursor.execute(
                "INSERT INTO auto_search_log (niche, leads_found) VALUES (%s,%s)",
                (niche, saved_count)
            )
            conn.commit()
            cursor.close()
            conn.close()

            # Send notification
            if saved_count > 0:
                add_notification(
                    f"🤖 Auto Search Found {saved_count} New Leads!",
                    f"Found {saved_count} new leads for '{niche}' in {country}. Check your Lead Bank!",
                    "success"
                )
                # Send email to team
                saved_leads_list = []
                conn2 = get_db()
                cur2 = conn2.cursor(dictionary=True)
                cur2.execute("SELECT * FROM leads ORDER BY created_at DESC LIMIT %s", (saved_count,))
                saved_leads_list = cur2.fetchall()
                cur2.close()
                conn2.close()
                send_new_leads_email(saved_leads_list, niche, country)
                print(f"✅ Auto-search saved {saved_count} leads for {niche} in {country}")
            else:
                print(f"ℹ️ Auto-search found no new leads for {niche} in {country}")

        except Exception as e:
            print(f"❌ Auto-search error: {e}")
            add_notification("⚠️ Auto Search Error", str(e), "error")
            threading.Event().wait(60 * 60)  # Wait 1 hour on error



@app.route("/api/auto-search/run", methods=["POST"])
def manual_auto_search():
    """Admin can trigger auto-search manually"""
    try:
        api_key = os.getenv("ANTHROPIC_API_KEY", "").strip()
        if not api_key or not api_key.startswith("sk-ant-"):
            return jsonify({"error": "API key not configured"}), 500

        data = request.get_json() or {}
        niche = data.get("niche", "Organic Apparel / Clothing")
        country = data.get("country", "India")

        client = anthropic.Anthropic(api_key=api_key)
        prompt = f"""You are a lead generation expert for Silasya (organic B2C) and Shoumitra (B2B export).
Find 5 fresh, realistic leads for:
Niche: {niche}
Country: {country}

Return ONLY a valid JSON array with 5 leads. Each must have:
- name, type (b2c/b2b), category, country, city, email, phone, website, description, why_good, potential_value, score (1-100), tags (array), source
Return ONLY the JSON array."""

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=3000,
            messages=[{"role": "user", "content": prompt}]
        )

        raw = message.content[0].text.strip()
        raw = raw.replace("```json", "").replace("```", "").strip()
        start = raw.find("[")
        end = raw.rfind("]")
        if start != -1 and end != -1:
            raw = raw[start:end+1]

        leads = json.loads(raw)
        saved_count = 0
        saved_leads = []

        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        for lead in leads:
            name = lead.get("name") or "Unknown"
            email = lead.get("email", "")
            country_val = lead.get("country", "")

            if email:
                cursor.execute("SELECT id FROM leads WHERE email=%s", (email,))
                if cursor.fetchone():
                    continue
            cursor.execute("SELECT id FROM leads WHERE name=%s AND country=%s", (name, country_val))
            if cursor.fetchone():
                continue

            tags = lead.get("tags", [])
            if isinstance(tags, list):
                tags = json.dumps(tags)

            cursor.execute("""
                INSERT INTO leads (name, type, category, country, city, email, phone,
                    website, description, why_good, potential_value, tags, score,
                    status, source, saved_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                name, lead.get("type","b2b"), lead.get("category",""),
                country_val, lead.get("city",""), email,
                lead.get("phone",""), lead.get("website",""),
                lead.get("description",""), lead.get("why_good",""),
                lead.get("potential_value",""), tags,
                int(lead.get("score",50)), "new",
                lead.get("source","Manual Search"), "Admin"
            ))
            saved_count += 1
            lead["id"] = cursor.lastrowid
            saved_leads.append(lead)

        conn.commit()
        cursor.close()
        conn.close()

        if saved_count > 0:
            add_notification(
                f"▶️ Manual Search Found {saved_count} New Leads!",
                f"Admin triggered search found {saved_count} new leads for '{niche}' in {country}.",
                "success"
            )
            send_new_leads_email(saved_leads, niche, country)

        return jsonify({"success": True, "saved": saved_count, "leads": saved_leads})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/test-email", methods=["POST"])
def test_email():
    try:
        sender = os.getenv("MAIL_SENDER", "").strip()
        password = os.getenv("MAIL_PASSWORD", "").strip()
        members = get_team_emails()
        
        if not sender:
            return jsonify({"error": "MAIL_SENDER not set in Railway Variables"})
        if not password:
            return jsonify({"error": "MAIL_PASSWORD not set in Railway Variables"})
        if not members:
            return jsonify({"error": "No team members found in Team tab"})
            
        import smtplib
        # Test SMTP connection directly
        try:
            server = smtplib.SMTP("smtp.gmail.com", 587, timeout=10)
            server.starttls()
            server.login(sender, password)
            server.quit()
            smtp_ok = True
            smtp_error = None
        except Exception as smtp_e:
            smtp_ok = False
            smtp_error = str(smtp_e)
            
        if smtp_ok:
            result = send_email_notification(
                "🧪 Test Email from Silasya Lead Finder",
                "<h2>Test email working!</h2><p>Hi {{name}}, your email notifications are set up correctly.</p>"
            )
            return jsonify({
                "success": result,
                "smtp": "connected",
                "sender": sender,
                "recipients": [m["email"] for m in members],
                "message": "Email sent!" if result else "Email failed"
            })
        else:
            return jsonify({
                "success": False,
                "smtp": "failed",
                "smtp_error": smtp_error,
                "sender": sender,
                "hint": "Check RESEND_API_KEY in Railway Variables - must start with re_"
            })
    except Exception as e:
        return jsonify({"error": str(e)})

# ─── Main ────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("🚀 Starting Silasya & Shoumitra Lead Finder...")
    print("📦 Initializing database...")
    init_db()
    port = int(os.getenv("PORT", 5000))
    debug = os.getenv("FLASK_DEBUG", "True").lower() == "true"
    print(f"✅ App running at http://localhost:{port}")
    print(f"📋 Share with team: http://YOUR_IP:{port}")
    # Start auto-search background thread
    auto_thread = threading.Thread(target=run_auto_search, daemon=True)
    auto_thread.start()
    print("🤖 Auto-search started (every 6 hours)")
    app.run(host="0.0.0.0", port=port, debug=debug)
